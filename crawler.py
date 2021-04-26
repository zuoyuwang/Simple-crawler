import requests
import bs4
from bs4 import BeautifulSoup
from requests.exceptions import RequestException
import openpyxl
from selenium import webdriver
import time


LINK1 = "https://www.linkedin.com/jobs/search/?f_TP=1%2C2&f_TPR=r604800&geoId=100025096&location=Toronto%2C%20Ontario%2C%20Canada&sortBy=DD"
DATE = '2020.08.31'
TOTAL = []
START = 1
END = 40


def get_normal_page(url, headers):
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            soup = BeautifulSoup(response.text, 'html.parser')
            return soup
        return None
    except RequestException:
        return None


def parse_one_page(soup, url):
    try:
        # shi = soup.find(class_='artdeco-inline-feedback__message')
        # if shi is None:
        #     pass
        # else:
        #     raise Exception
        result = []

        # title
        a = soup.find(class_='jobs-details-top-card__job-title').get_text().strip()
        if a in TOTAL:
            raise Exception
        TOTAL.append(a)
        result.append(a)

        # company and location
        b = soup.find(class_='jobs-details-top-card__company-info').get_text().strip().replace(' ','').split('\n')
        l = ['CompanyName', '', 'CompanyLocation']
        final = ''
        for i in b:
            if i not in l:
                final += i
                final += ' / '

        result.append(final)

        # location
        # c = soup.find(class_='jobs-details-top-card__exact-location').get_text().strip()
        # result.append(c)

        dic = {'Seniority Level':'N/A', 'Employment Type': 'N/A', 'Job Functions': 'N/A', 'Industry': 'N/A'}


        tags = soup.find_all(class_='jobs-box__group')
        for tag in tags:
           head = tag.find(class_='t-14 t-bold')
           if head is not None:
               head = head.get_text().strip()
           # s and e
           if head == 'Seniority Level' or head == 'Employment Type':
               top = tag.find(class_="t-14 mb3").get_text().strip()
               dic[head] = top

           elif head == 'Job Functions' or head == 'Industry':
               first = tag.find(class_='jobs-description-details__list-item t-14')
               text = first.get_text().strip()
               for fol in first.find_next_siblings(class_='jobs-description-details__list-item t-14'):
                   text = text + ', ' + fol.get_text().strip()

               dic[head] = text

        result.append(dic['Seniority Level'])
        result.append(dic['Employment Type'])
        result.append(dic['Job Functions'])
        result.append(dic['Industry'])


        result.append(DATE)
        result.append(url)

        return result
    except:
        pass

def save_to_excel(result):

    wb = openpyxl.load_workbook('LInkIn_Jobs.xlsx')
    ws = wb.active
    # ws['A1'] = 'Job-title'
    # ws['B1'] = 'Company'
    # ws['C1'] = 'Location'
    # ws['D1'] = 'Seniority level'
    # ws['E1'] = 'Employment type'
    # ws['F1'] = 'Job function'
    # ws['G1'] = 'Industries'
    # ws['H1'] = 'Link'
    # ws['I1'] = 'Date'
    ws.append(result)
    wb.save('LInkIn_Jobs.xlsx')


def main(link):
    headers = {
        'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_11_4) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/53.0.2785.116 Safari/537.36',
    }
    url = link
    # "https://www.linkedin.com/jobs/search/?f_TP=1&f_TPR=r86400&location=toronto&start=0"
    # 'https://www.linkedin.com/jobs/search/?f_TPR=r86400&location=toronto&redirect=false&position=1&pageNum=0&f_TP=1'

    driver = webdriver.Firefox()
    driver.get(url)

    time.sleep(1)

    a = driver.find_element_by_xpath(
        "//a[@class='nav__button-secondary']")
    a.click()

    time.sleep(1)
    driver.find_element_by_xpath("//input[@id='username']").send_keys(
        'tonywzy0731@gmail.com')
    driver.find_element_by_xpath("//input[@id='password']").send_keys(
        '13390943633Wzy.')
    driver.find_element_by_xpath(
        "//button[@class='btn__primary--large from__button--floating mercado-button--primary']").click()
    time.sleep(1.5)
    # num = START

    if START != 1:

        for j in range(9, START, 3):
            target = driver.find_elements_by_class_name("occludable-update")
            for i in range(2, len(target), 2):
                driver.execute_script('arguments[0].scrollIntoView();',
                                      target[i])
                time.sleep(0.5)

            no = driver.find_element_by_xpath(
                    "//button[@aria-label='Page {}']".format(j))

            no.click()

            time.sleep(1.5)

        time.sleep(2)


    for j in range(START, END + 1):

        target = driver.find_elements_by_class_name("occludable-update")
        for i in range(2, len(target), 2):
            driver.execute_script('arguments[0].scrollIntoView();', target[i])
            time.sleep(0.7)

        title = driver.find_elements_by_class_name("artdeco-entity-lockup__title")
        data = driver.page_source
        soup = BeautifulSoup(data, 'html.parser')
        web = soup.find_all(class_="job-card-list__title")

        for i in range(len(title)):
            title[i].click()
            time.sleep(1)
            data = driver.page_source
            soup = BeautifulSoup(data, 'html.parser')
            new_url = 'https://www.linkedin.com' + web[i]['href']
            lst = parse_one_page(soup, new_url)
            if lst is not None:
                save_to_excel(lst)

        # data = driver.page_source
        # soup = BeautifulSoup(data, 'html.parser')
        # web = soup.find_all(class_="job-card-list__title")
        #
        # for each in web:
        #     new_url = 'https://www.linkedin.com' + each['href']
        #     soup1 = get_normal_page(new_url, headers)
        #     if soup1 is None:
        #         pass
        #     else:
        #         result = []
        #         lst = parse_one_page(soup1, new_url)
        #         if lst is not None:
        #             result.append(lst)
        #         save_to_excel(result)

        no = None
        if j == END:
            driver.quit()
            break
        else:
            no = driver.find_element_by_xpath(
                "//button[@aria-label='Page {}']".format(j + 1))

        no.click()
        time.sleep(2)


if __name__ == '__main__':
    main(LINK1)
